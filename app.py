import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import psycopg2
import psycopg2.extras
from datetime import datetime, date
from zoneinfo import ZoneInfo
import io

# Malaysia timezone (UTC+8)
MY_TZ = ZoneInfo("Asia/Kuala_Lumpur")

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Solar PV Database & Calculator",
    page_icon="☀️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Custom CSS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Orbitron:wght@400;700;900&family=Exo+2:wght@300;400;600;700&display=swap');

:root {
    --gold: #FFB800; --orange: #FF6B00; --cyan: #00D4FF;
    --navy: #0A0E1A; --panel: #0D1525; --card: #111827;
    --green: #00E676; --red: #FF4444;
    --txt: #E8EAF0; --muted: #8892A4;
}
html, body, [class*="css"] { font-family: 'Exo 2', sans-serif; }
.stApp { background: linear-gradient(135deg, #0A0E1A 0%, #0D1525 60%, #0A1628 100%); color: var(--txt); }

.solar-header {
    background: linear-gradient(90deg,#0A0E1A,#1a2744,#0A0E1A);
    border-bottom: 2px solid var(--gold);
    padding: 1.4rem 2rem; margin: -1rem -1rem 1.5rem -1rem;
    text-align: center; position: relative; overflow: hidden;
}
.solar-header::before {
    content:''; position:absolute; top:0;left:0;right:0;bottom:0;
    background:radial-gradient(ellipse at center top,rgba(255,184,0,.12) 0%,transparent 70%);
}
.solar-header h1 { font-family:'Orbitron',monospace; font-size:2rem; font-weight:900;
    color:var(--gold); letter-spacing:4px; text-shadow:0 0 30px rgba(255,184,0,.5); margin:0; }
.solar-header p { color:var(--muted); font-size:.9rem; letter-spacing:2px; margin:.3rem 0 0; }

.kpi-card { background:linear-gradient(135deg,#111827,#1a2744);
    border:1px solid rgba(255,184,0,.2); border-radius:12px;
    padding:1rem; text-align:center; margin-bottom:.8rem;
    transition:all .3s; box-shadow:0 4px 15px rgba(0,0,0,.3); }
.kpi-card:hover { border-color:var(--gold); box-shadow:0 4px 25px rgba(255,184,0,.15); transform:translateY(-2px); }
.kpi-val { font-family:'Orbitron',monospace; font-size:1.6rem; font-weight:700; color:var(--gold); }
.kpi-lbl { font-size:.8rem; color:var(--muted); letter-spacing:1px; text-transform:uppercase; }
.kpi-unit { font-size:.72rem; color:var(--cyan); }

.sec { font-family:'Orbitron',monospace; font-size:1rem; color:var(--gold);
    letter-spacing:2px; border-left:3px solid var(--gold); padding-left:.8rem; margin:1.2rem 0 .8rem; }

.res-box { border-radius:10px; padding:1rem 1.2rem; margin:.4rem 0; }
.res-measured { background:rgba(0,212,255,.08); border:1px solid rgba(0,212,255,.35); }
.res-calculated { background:rgba(255,184,0,.08); border:1px solid rgba(255,184,0,.35); }
.res-diff-ok { background:rgba(0,230,118,.08); border:1px solid rgba(0,230,118,.35); }
.res-diff-warn { background:rgba(255,68,68,.08); border:1px solid rgba(255,68,68,.35); }
.res-title { font-family:'Orbitron',monospace; font-size:.78rem; letter-spacing:1px; margin-bottom:.4rem; }
.res-val { font-size:1.5rem; font-weight:700; font-family:'Orbitron',monospace; }
.res-formula { font-size:.75rem; color:var(--muted); margin-top:.3rem; font-style:italic; }

.factor-row { display:flex; justify-content:space-between; padding:.35rem .6rem;
    border-bottom:1px solid rgba(255,184,0,.08); font-size:.85rem; }
.factor-row:last-child { border-bottom:none; }
.factor-name { color:var(--muted); }
.factor-val { font-family:'Orbitron',monospace; color:var(--gold); font-size:.8rem; }

.badge-online { display:inline-block; padding:.2rem .8rem; border-radius:20px;
    font-size:.72rem; font-weight:600; letter-spacing:1px;
    background:rgba(0,230,118,.15); color:#00E676; border:1px solid #00E676; }

[data-testid="stSidebar"] { background:linear-gradient(180deg,#0A0E1A,#0D1525); border-right:1px solid rgba(255,184,0,.2); }

.stButton>button { background:linear-gradient(135deg,var(--orange),var(--gold));
    color:#0A0E1A; font-family:'Orbitron',monospace; font-weight:700; letter-spacing:1px;
    border:none; border-radius:8px; padding:.5rem 1.2rem;
    transition:all .3s; box-shadow:0 4px 15px rgba(255,107,0,.3); }
.stButton>button:hover { box-shadow:0 4px 25px rgba(255,184,0,.5); transform:translateY(-1px); }

.stTextInput>div>div>input, .stNumberInput>div>div>input,
.stSelectbox>div>div, .stDateInput>div>div>input {
    background:#111827!important; border:1px solid rgba(255,184,0,.25)!important;
    color:var(--txt)!important; border-radius:8px!important; }

.stTabs [data-baseweb="tab-list"] { background:transparent; border-bottom:1px solid rgba(255,184,0,.2); gap:.3rem; }
.stTabs [data-baseweb="tab"] { font-family:'Orbitron',monospace; font-size:.7rem; letter-spacing:1px;
    color:var(--muted)!important; background:transparent!important;
    border-radius:6px 6px 0 0!important; padding:.5rem 1rem!important;
    border:1px solid transparent!important; border-bottom:none!important; }
.stTabs [aria-selected="true"] { color:var(--gold)!important;
    background:rgba(255,184,0,.08)!important; border-color:rgba(255,184,0,.3)!important; }

.streamlit-expanderHeader { font-family:'Orbitron',monospace; font-size:.8rem; color:var(--gold)!important; }
hr { border-color:rgba(255,184,0,.15); }
</style>
""", unsafe_allow_html=True)


# ── Database (PostgreSQL via Supabase — persistent storage) ────────────────────
def get_conn():
    return psycopg2.connect(st.secrets["DATABASE_URL"], sslmode="prefer")

def init_db():
    with get_conn() as c:
        cur = c.cursor()
        cur.execute("""CREATE TABLE IF NOT EXISTS measured_data (
            id SERIAL PRIMARY KEY,
            recorded_at TIMESTAMP NOT NULL,
            site_name TEXT DEFAULT 'Default',
            irradiance REAL, pv_temp REAL, ambient_temp REAL,
            voc_measured REAL, vmp_measured REAL,
            isc_measured REAL, imp_measured REAL,
            power_measured REAL,
            notes TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS calc_results (
            id SERIAL PRIMARY KEY,
            measured_id INTEGER,
            calc_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            site_name TEXT,
            p_max_stc REAL, voc_stc REAL, vmp_stc REAL,
            isc_stc REAL, imp_stc REAL,
            gamma_pmax REAL, beta_voc REAL, alpha_isc REAL,
            noct REAL, t_stc REAL,
            irradiance REAL, t_mod REAL, dirt_pct REAL,
            shading_pct REAL, f_lid REAL, f_age REAL,
            f_mm REAL, bifacial_gain REAL,
            f_degrad REAL, f_temp_p REAL, f_g REAL,
            f_clean REAL, f_unshade REAL,
            f_temp_i REAL, f_temp_v REAL,
            p_max_calc REAL, isc_calc REAL, imp_calc REAL,
            voc_calc REAL, vmp_calc REAL,
            p_max_meas REAL, isc_meas REAL, imp_meas REAL,
            voc_meas REAL, vmp_meas REAL
        )""")
        cur.execute("""CREATE TABLE IF NOT EXISTS sites (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL, location TEXT, capacity REAL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""INSERT INTO sites(name,location,capacity)
            VALUES('Default Site','Malaysia',1.0)
            ON CONFLICT(name) DO NOTHING""")
        c.commit()

def load_measured(site=None, start=None, end=None):
    q = "SELECT * FROM measured_data WHERE 1=1"
    p = []
    if site and site != "All Sites": q += " AND site_name=%s"; p.append(site)
    if start: q += " AND DATE(recorded_at)>=%s"; p.append(str(start))
    if end:   q += " AND DATE(recorded_at)<=%s"; p.append(str(end))
    q += " ORDER BY recorded_at DESC"
    with get_conn() as c:
        df = pd.read_sql_query(q, c, params=p)
    if not df.empty:
        df["recorded_at"] = pd.to_datetime(df["recorded_at"], utc=True).dt.tz_convert("Asia/Kuala_Lumpur").dt.tz_localize(None)
    return df

def load_calcs(site=None, start=None, end=None):
    q = "SELECT * FROM calc_results WHERE 1=1"
    p = []
    if site and site != "All Sites": q += " AND site_name=%s"; p.append(site)
    if start: q += " AND DATE(calc_at)>=%s"; p.append(str(start))
    if end:   q += " AND DATE(calc_at)<=%s"; p.append(str(end))
    q += " ORDER BY calc_at DESC"
    with get_conn() as c:
        df = pd.read_sql_query(q, c, params=p)
    if not df.empty:
        df["calc_at"] = pd.to_datetime(df["calc_at"], utc=True).dt.tz_convert("Asia/Kuala_Lumpur").dt.tz_localize(None)
    return df

def load_sites():
    with get_conn() as c:
        return pd.read_sql_query("SELECT name FROM sites ORDER BY name", c)["name"].tolist()

def insert_measured(d):
    cols = ", ".join(d.keys())
    ph   = ", ".join(["%s"] * len(d))
    sql  = f"INSERT INTO measured_data ({cols}) VALUES ({ph}) RETURNING id"
    with get_conn() as c:
        cur = c.cursor()
        cur.execute(sql, list(d.values()))
        new_id = cur.fetchone()[0]
        c.commit()
        return new_id

def insert_calc(d):
    cols = ", ".join(d.keys())
    ph   = ", ".join(["%s"] * len(d))
    sql  = f"INSERT INTO calc_results ({cols}) VALUES ({ph})"
    with get_conn() as c:
        cur = c.cursor()
        cur.execute(sql, list(d.values()))
        c.commit()

def delete_measured(rid):
    with get_conn() as c:
        cur = c.cursor()
        cur.execute("DELETE FROM calc_results WHERE measured_id=%s", (rid,))
        cur.execute("DELETE FROM measured_data WHERE id=%s", (rid,))
        c.commit()

init_db()


# ═══════════════════════════════════════════════════════════════════════════════
#  SEDA COMPUTATION ENGINE  (Fundamentals of Solar PV Technology, Ch 5.7–5.8)
# ═══════════════════════════════════════════════════════════════════════════════
def calc_factors(irradiance, t_mod, t_stc,
                 gamma_pmax, beta_voc, alpha_isc,
                 dirt_pct, shading_pct, f_lid, f_age, f_mm):
    f_g       = irradiance / 1000                                    # Eq 5.9
    f_degrad  = f_lid * f_age                                        # Eq 5.8
    f_temp_p  = 1 + (gamma_pmax / 100) * (t_mod - t_stc)            # Eq 5.10
    f_temp_i  = 1 + (alpha_isc  / 100) * (t_mod - t_stc)            # Eq 5.12
    f_temp_v  = 1 + (beta_voc   / 100) * (t_mod - t_stc)            # Eq 5.14
    f_clean   = (100 - dirt_pct)    / 100
    f_unshade = (100 - shading_pct) / 100
    return dict(f_g=f_g, f_degrad=f_degrad, f_temp_p=f_temp_p,
                f_temp_i=f_temp_i, f_temp_v=f_temp_v,
                f_clean=f_clean, f_unshade=f_unshade, f_mm=f_mm,
                f_lid=f_lid, f_age=f_age)

def calc_outputs(p_stc, voc_stc, vmp_stc, isc_stc, imp_stc, factors,
                 bifacial_gain=0, n_s=1, n_p=1):
    F = factors
    p_max = (p_stc * (1 + bifacial_gain/100)
             * F["f_mm"] * F["f_degrad"] * F["f_temp_p"]
             * F["f_g"] * F["f_clean"] * F["f_unshade"]
             * n_s * n_p)                                            # Eq 5.7
    isc   = isc_stc * F["f_temp_i"] * F["f_g"] * F["f_clean"] * F["f_unshade"] * n_p  # Eq 5.11
    imp   = imp_stc * F["f_temp_i"] * F["f_g"] * F["f_clean"] * F["f_unshade"] * n_p
    voc   = voc_stc * F["f_temp_v"] * n_s                           # Eq 5.13
    vmp   = vmp_stc * F["f_temp_v"] * n_s
    return dict(p_max=p_max, isc=isc, imp=imp, voc=voc, vmp=vmp)

def malaysia_pv_temp(irradiance, t_amb):
    """Eq 5.6 — Malaysian climate empirical model (Zainuddin, H., 2014)."""
    return -15.76 + 0.02 * irradiance + 1.64 * t_amb

def noct_pv_temp(irradiance, t_amb, noct):
    """Eq 5.4 — NOCT temperature model."""
    return t_amb + irradiance * (noct - 20) / 800

def pct_diff(calc, meas):
    if meas and meas != 0:
        return (calc - meas) / abs(meas) * 100
    return None


# ── Header ──────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="solar-header">
  <h1>☀ SOLAR PV DATABASE & CALCULATOR</h1>
  <p>FIELD MEASURED vs CALCULATED · SEDA FUNDAMENTALS OF SOLAR PV TECHNOLOGY</p>
</div>
""", unsafe_allow_html=True)


# ── Sidebar ──────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown('<p class="sec">⚙ NAVIGATION</p>', unsafe_allow_html=True)
    page = st.radio("Module", [
        "📊 Dashboard",
        "🔬 Field Measurement",
        "🧮 SEDA Calculator",
        "📐 Comparison Analysis",
        "📤 Import / Export",
        "📈 Analytics",
    ], label_visibility="collapsed")
    st.markdown("---")
    st.markdown('<p class="sec">🔍 FILTER</p>', unsafe_allow_html=True)
    sites     = ["All Sites"] + load_sites()
    sel_site  = st.selectbox("Site", sites)
    
    # Get current time in Malaysia timezone to find today's date safely
    now_my = datetime.now(MY_TZ)
    today_my = now_my.date()
    
    sel_start = st.date_input("From", value=date(2024, 1, 1))
    sel_end   = st.date_input("To",   value=today_my)
    st.markdown("---")
    st.markdown('<span class="badge-online">● SYSTEM ONLINE</span>', unsafe_allow_html=True)
    st.caption(f"v3.0 · {now_my.strftime('%d %b %Y  %H:%M')}")

df_meas = load_measured(sel_site, sel_start, sel_end)
df_calc = load_calcs(sel_site, sel_start, sel_end)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
if page == "📊 Dashboard":
    st.markdown('<p class="sec">📊 SYSTEM OVERVIEW</p>', unsafe_allow_html=True)

    def kpi(col, icon, val, lbl, unit):
        col.markdown(f"""<div class="kpi-card">
            <div style="font-size:1.3rem">{icon}</div>
            <div class="kpi-val">{val}</div>
            <div class="kpi-lbl">{lbl}</div>
            <div class="kpi-unit">{unit}</div></div>""")

    k1,k2,k3,k4,k5,k6 = st.columns(6)
    if df_meas.empty:
        for c,i,v,l,u in zip([k1,k2,k3,k4,k5,k6],
            ["📋","🌞","⚡","🌡️","🧮","📐"],
            ["0","—","—","—","0","—"],
            ["Field Records","Avg Irradiance","Max Power","Avg PV Temp","Computations","Avg Efficiency"],
            ["","W/m²","W","°C","","W·m²/W"]):
            kpi(c,i,v,l,u)
    else:
        eff = (df_meas["power_measured"] / df_meas["irradiance"].replace(0, np.nan)).mean()
        kpi(k1,"📋", len(df_meas),                               "Field Records",   "")
        kpi(k2,"🌞", f"{df_meas['irradiance'].mean():.0f}",      "Avg Irradiance",  "W/m²")
        kpi(k3,"⚡", f"{df_meas['power_measured'].max():.1f}",   "Max Power",       "W")
        kpi(k4,"🌡️", f"{df_meas['pv_temp'].mean():.1f}",        "Avg PV Temp",     "°C")
        kpi(k5,"🧮", len(df_calc),                               "Computations",    "")
        kpi(k6,"📐", f"{eff:.3f}",                               "Avg Efficiency",  "W·m²/W")

    if not df_meas.empty:
        st.markdown("---")
        st.markdown('<p class="sec">📈 TIME-SERIES — FIELD DATA</p>', unsafe_allow_html=True)
        cd = df_meas.sort_values("recorded_at")
        fig = make_subplots(rows=2, cols=2,
            subplot_titles=["Solar Irradiance (W/m²)", "Measured Power (W)",
                            "PV & Ambient Temperature (°C)", "Voc & Vmp (V)"],
            vertical_spacing=0.14)
        kw = dict(mode="lines+markers", marker=dict(size=4))
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.irradiance,      name="Irradiance",   line=dict(color="#FFB800",width=2),**kw), row=1, col=1)
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.power_measured,  name="Power",        line=dict(color="#FF6B00",width=2),**kw), row=1, col=2)
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.pv_temp,         name="PV Temp",      line=dict(color="#00D4FF",width=2),**kw), row=2, col=1)
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.ambient_temp,    name="Ambient Temp", line=dict(color="#00E676",width=2),**kw), row=2, col=1)
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.voc_measured,    name="Voc",          line=dict(color="#FF6BFF",width=2),**kw), row=2, col=2)
        fig.add_trace(go.Scatter(x=cd.recorded_at, y=cd.vmp_measured,    name="Vmp",          line=dict(color="#FFB8FF",width=2,dash="dash"),**kw), row=2, col=2)
        fig.update_layout(height=520, paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(13,21,37,0.8)", font=dict(color="#8892A4",family="Exo 2"),
            showlegend=True, legend=dict(bgcolor="rgba(0,0,0,.5)"))
        for ann in fig.layout.annotations: ann.font.color = "#FFB800"
        fig.update_xaxes(gridcolor="rgba(255,184,0,.08)")
        fig.update_yaxes(gridcolor="rgba(255,184,0,.08)")
        st.plotly_chart(fig, use_container_width=True)

    if not df_calc.empty:
        st.markdown('<p class="sec">⚖ MEASURED vs CALCULATED POWER</p>', unsafe_allow_html=True)
        fig2 = go.Figure()
        fig2.add_trace(go.Scatter(x=df_calc.calc_at, y=df_calc.p_max_calc,
            name="Calculated (SEDA)", line=dict(color="#FFB800", width=2.5)))
        fig2.add_trace(go.Scatter(x=df_calc.calc_at, y=df_calc.p_max_meas,
            name="Measured (Field)",  line=dict(color="#00D4FF", width=2.5, dash="dot")))
        fig2.update_layout(height=320, paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(13,21,37,0.8)", font=dict(color="#8892A4",family="Exo 2"),
            legend=dict(bgcolor="rgba(0,0,0,.5)"),
            title="Power Output: SEDA Calculated vs Field Measured",
            title_font=dict(color="#FFB800"))
        fig2.update_xaxes(gridcolor="rgba(255,184,0,.08)")
        fig2.update_yaxes(gridcolor="rgba(255,184,0,.08)", title="Power (W)")
        st.plotly_chart(fig2, use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: FIELD MEASUREMENT
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🔬 Field Measurement":
    st.markdown('<p class="sec">🔬 RECORD FIELD MEASUREMENT (ROC)</p>', unsafe_allow_html=True)
    st.caption("Enter values measured directly at the site under Real Operating Conditions (ROC).")

    # Get target timezone context safely before initializing standard widgets
    now_my = datetime.now(MY_TZ)
    today_my = now_my.date()
    time_my = now_my.time()

    with st.form("meas_form", clear_on_submit=True):
        c1, c2, c3 = st.columns(3)
        with c1:
            rec_date = st.date_input("📅 Date", value=today_my)
            rec_time = st.time_input("🕐 Time", value=time_my)
        with c2:
            site_name = st.selectbox("🏭 Site", load_sites())
        with c3:
            notes = st.text_input("📝 Notes (optional)")

        st.markdown("**🌞 Environmental Parameters**")
        e1, e2, e3 = st.columns(3)
        with e1: irr  = st.number_input("Solar Irradiance (W/m²)",  0.0, 1500.0, 800.0, 1.0)
        with e2: pvt  = st.number_input("PV Temperature (°C)",      -20.0, 100.0, 45.0, 0.1)
        with e3: ambt = st.number_input("Ambient Temperature (°C)", -20.0,  60.0, 30.0, 0.1)

        st.markdown("**⚡ Electrical Parameters — Field Measured Values**")
        p1,p2,p3,p4,p5 = st.columns(5)
        with p1: voc = st.number_input("Voc (V)",    0.0, 5000.0,  40.5, 0.01)
        with p2: vmp = st.number_input("Vmp (V)",    0.0, 5000.0,  33.2, 0.01)
        with p3: isc = st.number_input("Isc (A)",    0.0,  500.0, 10.05, 0.01)
        with p4: imp = st.number_input("Imp (A)",    0.0,  500.0,  9.49, 0.01)
        with p5: pwr = st.number_input("Power (W)",  0.0, 1e7,    315.0, 0.1)

        submitted = st.form_submit_button("💾 SAVE MEASUREMENT", use_container_width=True)

    if submitted:
        # Combine date + time, mark as Malaysia time (UTC+8)
        ts = datetime.combine(rec_date, rec_time, tzinfo=MY_TZ).strftime("%Y-%m-%d %H:%M:%S")
        mid = insert_measured(dict(recorded_at=ts, site_name=site_name,
            irradiance=irr, pv_temp=pvt, ambient_temp=ambt,
            voc_measured=voc, vmp_measured=vmp,
            isc_measured=isc, imp_measured=imp,
            power_measured=pwr, notes=notes))
        st.success(f"✅ Field measurement saved — Record ID #{mid}")

    st.markdown("---")
    st.markdown('<p class="sec">🏭 MANAGE SITES</p>', unsafe_allow_html=True)
    with st.form("site_form", clear_on_submit=True):
        sc1, sc2 = st.columns(2)
        with sc1: new_site = st.text_input("Site Name")
        with sc2: location = st.text_input("Location")
        if st.form_submit_button("➕ Add Site", use_container_width=True):
            if new_site:
                try:
                    with get_conn() as conn:
                        cur = conn.cursor()
                        cur.execute(
                            "INSERT INTO sites(name,location,capacity) VALUES(%s,%s,%s)",
                            (new_site, location, None))
                        conn.commit()
                    st.success(f"Site '{new_site}' added successfully!")
                except Exception:
                    st.error("A site with this name already exists.")

    if not df_meas.empty:
        st.markdown("---")
        st.markdown('<p class="sec">📋 RECENT FIELD RECORDS</p>', unsafe_allow_html=True)
        show_cols = ["id","recorded_at","site_name","irradiance","pv_temp","ambient_temp",
                     "voc_measured","vmp_measured","isc_measured","imp_measured","power_measured"]
        rename_map = {
            "irradiance":"Irr (W/m²)", "pv_temp":"PV Temp (°C)", "ambient_temp":"Amb Temp (°C)",
            "voc_measured":"Voc (V)",  "vmp_measured":"Vmp (V)",
            "isc_measured":"Isc (A)",  "imp_measured":"Imp (A)",
            "power_measured":"Power (W)", "recorded_at":"Timestamp", "site_name":"Site",
        }
        st.dataframe(df_meas[show_cols].rename(columns=rename_map).head(30),
                     use_container_width=True, height=350)
        st.markdown("**Delete a record by ID:**")
        del_id = st.number_input("Record ID to delete", min_value=1, step=1)
        if st.button("🗑️ Delete Record"):
            delete_measured(int(del_id))
            st.success("Record deleted successfully.")
            st.rerun()


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: SEDA CALCULATOR
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "🧮 SEDA Calculator":
    st.markdown('<p class="sec">🧮 SEDA COMPUTATION — ROC OUTPUT ESTIMATION</p>', unsafe_allow_html=True)
    st.caption("Based on **Fundamentals of Solar Photovoltaics Technology (SEDA Malaysia)** — Equations 5.4, 5.6 – 5.14")

    tab_mod, tab_arr = st.tabs(["📦 MODULE / STRING", "🏭 ARRAY GUIDE"])

    # ── MODULE TAB ─────────────────────────────────────────────────────────────
    with tab_mod:
        st.markdown("**1️⃣ PV Module Datasheet (Standard Test Conditions — STC)**")
        d1, d2, d3 = st.columns(3)
        with d1:
            p_stc   = st.number_input("P_max STC (W)",   100.0, 1000.0, 315.0,  0.1)
            voc_stc = st.number_input("Voc STC (V)",       1.0,  200.0,  40.5,  0.01)
            vmp_stc = st.number_input("Vmp STC (V)",       1.0,  200.0,  33.2,  0.01)
        with d2:
            isc_stc = st.number_input("Isc STC (A)",       0.1,   50.0, 10.05,  0.01)
            imp_stc = st.number_input("Imp STC (A)",        0.1,   50.0,  9.49,  0.01)
            noct    = st.number_input("NOCT (°C)",         20.0,   60.0,  47.0,  0.1)
        with d3:
            gamma   = st.number_input("γ_Pmax (%/°C)",    -1.0,    0.0, -0.37,  0.001)
            beta_v  = st.number_input("β_Voc  (%/°C)",    -1.0,    0.0, -0.28,  0.001)
            alpha_i = st.number_input("α_Isc  (%/°C)",     0.0,    1.0,  0.048, 0.001)
            t_stc   = st.number_input("T_STC (°C)",        0.0,   50.0,  25.0,  0.1)

        st.markdown("---")
        st.markdown("**2️⃣ Real Operating Conditions (ROC) at Site**")
        r1, r2, r3 = st.columns(3)
        with r1:
            irr_roc   = st.number_input("Irradiance (W/m²)",      0.0, 1500.0, 800.0, 1.0, key="irr_roc")
            t_mod_roc = st.number_input("T_mod / T_cell (°C)",  -20.0,  120.0,  65.0, 0.1, key="tmod")
            temp_mode = st.selectbox("PV Temperature Method", [
                "Manual — use T_mod above",
                "NOCT Model (Eq 5.4)",
                "Malaysian Climate Model (Eq 5.6)",
            ])
            t_amb_roc = st.number_input("Ambient Temperature (°C)", -20.0, 60.0, 30.0, 0.1, key="tamb")
        with r2:
            dirt_pct    = st.slider("Dirt / Soiling (%)",  0.0, 20.0,   2.0, 0.5)
            shading_pct = st.slider("Shading (%)",         0.0, 100.0,  0.0, 1.0)
            f_lid_val   = st.number_input("f_LID — Light Induced Degradation factor", 0.5, 1.0, 0.975, 0.001)
            f_age_val   = st.number_input("f_age — Aging factor",                     0.5, 1.0, 1.0,   0.001)
        with r3:
            f_mm_val  = st.number_input("f_mm — Mismatch factor", 0.5, 1.0, 1.0, 0.001)
            bifacial  = st.number_input("Bifacial Gain (%)",      0.0, 30.0, 0.0, 1.0)
            n_s       = st.number_input("N_s — Modules in series",   1, 100,  1,  1)
            n_p       = st.number_input("N_p — Strings in parallel", 1, 100,  1,  1)

        st.markdown("---")
        st.markdown("**3️⃣ Field Measured Values (for comparison — enter 0 to skip)**")
        m1,m2,m3,m4,m5 = st.columns(5)
        with m1: p_meas   = st.number_input("Measured Power (W)",   0.0, 1e7,    0.0, 0.1,  key="pm")
        with m2: voc_meas = st.number_input("Measured Voc (V)",     0.0, 5000.0, 0.0, 0.01, key="vm")
        with m3: vmp_meas = st.number_input("Measured Vmp (V)",     0.0, 5000.0, 0.0, 0.01, key="vmpm")
        with m4: isc_meas = st.number_input("Measured Isc (A)",     0.0,  500.0, 0.0, 0.01, key="im")
        with m5: imp_meas = st.number_input("Measured Imp (A)",     0.0,  500.0, 0.0, 0.01, key="impm")

        site_calc = st.selectbox("Site (for saving results)", load_sites(), key="site_calc")

        if st.button("⚡ CALCULATE NOW", use_container_width=True):
            # Determine T_mod
            if temp_mode == "NOCT Model (Eq 5.4)":
                t_mod_used = noct_pv_temp(irr_roc, t_amb_roc, noct)
                temp_note  = f"NOCT Model [Eq 5.4]: T_x = T_amb + [G×(NOCT−20)/800] = {t_mod_used:.2f} °C"
            elif temp_mode == "Malaysian Climate Model (Eq 5.6)":
                t_mod_used = malaysia_pv_temp(irr_roc, t_amb_roc)
                temp_note  = f"Malaysian Model [Eq 5.6]: T_x = −15.76 + 0.02G + 1.64×T_amb = {t_mod_used:.2f} °C"
            else:
                t_mod_used = t_mod_roc
                temp_note  = f"Manual input: T_mod = {t_mod_used:.2f} °C"

            F   = calc_factors(irr_roc, t_mod_used, t_stc,
                               gamma, beta_v, alpha_i,
                               dirt_pct, shading_pct, f_lid_val, f_age_val, f_mm_val)
            OUT = calc_outputs(p_stc, voc_stc, vmp_stc, isc_stc, imp_stc,
                               F, bifacial, n_s, n_p)

            st.session_state["last_calc"] = dict(
                F=F, OUT=OUT, t_mod=t_mod_used,
                irr=irr_roc, p_stc_base=p_stc*n_s*n_p,
                p_meas=p_meas, voc_meas=voc_meas, vmp_meas=vmp_meas,
                isc_meas=isc_meas, imp_meas=imp_meas,
                temp_note=temp_note,
                gamma=gamma, beta_v=beta_v, alpha_i=alpha_i,
                t_stc=t_stc, n_s=n_s, n_p=n_p,
            )

            insert_calc(dict(
                site_name=site_calc,
                p_max_stc=p_stc, voc_stc=voc_stc, vmp_stc=vmp_stc,
                isc_stc=isc_stc, imp_stc=imp_stc,
                gamma_pmax=gamma, beta_voc=beta_v, alpha_isc=alpha_i,
                noct=noct, t_stc=t_stc,
                irradiance=irr_roc, t_mod=t_mod_used,
                dirt_pct=dirt_pct, shading_pct=shading_pct,
                f_lid=f_lid_val, f_age=f_age_val, f_mm=f_mm_val,
                bifacial_gain=bifacial,
                f_degrad=F["f_degrad"], f_temp_p=F["f_temp_p"], f_g=F["f_g"],
                f_clean=F["f_clean"], f_unshade=F["f_unshade"],
                f_temp_i=F["f_temp_i"], f_temp_v=F["f_temp_v"],
                p_max_calc=OUT["p_max"], isc_calc=OUT["isc"],
                imp_calc=OUT["imp"], voc_calc=OUT["voc"], vmp_calc=OUT["vmp"],
                p_max_meas=p_meas, isc_meas=isc_meas, imp_meas=imp_meas,
                voc_meas=voc_meas, vmp_meas=vmp_meas,
                measured_id=None,
            ))

        # ── RESULTS ────────────────────────────────────────────────────────────
        if "last_calc" in st.session_state:
            R = st.session_state["last_calc"]
            F = R["F"]; OUT = R["OUT"]

            st.markdown("---")
            st.markdown('<p class="sec">📊 COMPUTATION RESULTS</p>', unsafe_allow_html=True)
            st.info(f"🌡️ {R['temp_note']}")

            with st.expander("🔢 Adjustment Factors — Step-by-Step"):
                factors_display = [
                    ("f_g  — Irradiance factor              [Eq 5.9]",   F["f_g"]),
                    ("f_LID — Light Induced Degradation factor",         F["f_lid"]),
                    ("f_age — Aging factor",                             F["f_age"]),
                    ("f_degrad = f_LID × f_age             [Eq 5.8]",   F["f_degrad"]),
                    ("f_temp_p — Temperature factor (Power) [Eq 5.10]", F["f_temp_p"]),
                    ("f_temp_i — Temperature factor (Current) [Eq 5.12]", F["f_temp_i"]),
                    ("f_temp_v — Temperature factor (Voltage) [Eq 5.14]", F["f_temp_v"]),
                    ("f_clean — Dirt / Soiling factor",                  F["f_clean"]),
                    ("f_unshade — Shading factor",                      F["f_unshade"]),
                    ("f_mm — Mismatch factor",                          F["f_mm"]),
                ]
                for name, val in factors_display:
                    st.markdown(f"""<div class="factor-row">
                        <span class="factor-name">{name}</span>
                        <span class="factor-val">{val:.4f}</span>
                    </div>""", unsafe_allow_html=True)

            st.markdown("**⚡ Results — Calculated (SEDA) vs Field Measured**")
            params = [
                ("POWER",       "P_max", "W", OUT["p_max"], R["p_meas"]),
                ("VOLTAGE OC",  "Voc",   "V", OUT["voc"],   R["voc_meas"]),
                ("VOLTAGE MP",  "Vmp",   "V", OUT["vmp"],   R["vmp_meas"]),
                ("CURRENT SC",  "Isc",   "A", OUT["isc"],   R["isc_meas"]),
                ("CURRENT MP",  "Imp",   "A", OUT["imp"],   R["imp_meas"]),
            ]
            for pname, sym, unit, calc_v, meas_v in params:
                col_c, col_m, col_d = st.columns([1, 1, 1])
                with col_c:
                    st.markdown(f"""<div class="res-box res-calculated">
                        <div class="res-title" style="color:#FFB800">🧮 CALCULATED — {pname}</div>
                        <div class="res-val" style="color:#FFB800">{calc_v:.3f} <span style="font-size:.9rem">{unit}</span></div>
                        <div class="res-formula">{sym}_calc — SEDA Formula</div>
                    </div>""", unsafe_allow_html=True)
                with col_m:
                    if meas_v and meas_v > 0:
                        st.markdown(f"""<div class="res-box res-measured">
                            <div class="res-title" style="color:#00D4FF">🔬 MEASURED — {pname}</div>
                            <div class="res-val" style="color:#00D4FF">{meas_v:.3f} <span style="font-size:.9rem">{unit}</span></div>
                            <div class="res-formula">{sym}_measured — Field / Site</div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        st.markdown(f"""<div class="res-box" style="border:1px solid rgba(136,146,164,.3);border-radius:10px;padding:1rem .8rem">
                            <div class="res-title" style="color:#8892A4">🔬 MEASURED — {pname}</div>
                            <div class="res-val" style="color:#8892A4">—</div>
                            <div class="res-formula">No measured value entered</div>
                        </div>""", unsafe_allow_html=True)
                with col_d:
                    if meas_v and meas_v > 0:
                        diff = pct_diff(calc_v, meas_v)
                        ok = abs(diff) <= 5
                        cls = "res-diff-ok" if ok else "res-diff-warn"
                        clr = "#00E676" if ok else "#FF4444"
                        icon = "✅" if ok else "⚠️"
                        msg = "Within ±5% tolerance" if ok else "Exceeds ±5% — investigate!"
                        st.markdown(f"""<div class="res-box {cls}">
                            <div class="res-title" style="color:{clr}">{icon} DIFFERENCE</div>
                            <div class="res-val" style="color:{clr}">{diff:+.2f}%</div>
                            <div class="res-formula">{msg}</div>
                        </div>""", unsafe_allow_html=True)
                    else:
                        st.markdown(f"""<div class="res-box" style="border:1px solid rgba(136,146,164,.3);border-radius:10px;padding:1rem .8rem">
                            <div class="res-title" style="color:#8892A4">📐 DIFFERENCE</div>
                            <div class="res-val" style="color:#8892A4">—</div>
                        </div>""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: COMPARISON ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📐 Comparison Analysis":
    st.markdown('<p class="sec">📐 COMPARISON ANALYSIS — CALCULATED vs MEASURED</p>', unsafe_allow_html=True)

    if df_calc.empty:
        st.info("No computation records found. Run a computation in the 🧮 SEDA Calculator first.")
    else:
        df_c = df_calc.copy()
        for param, c_col, m_col in [
            ("Power", "p_max_calc", "p_max_meas"),
            ("Voc",   "voc_calc",   "voc_meas"),
            ("Vmp",   "vmp_calc",   "vmp_meas"),
            ("Isc",   "isc_calc",   "isc_meas"),
            ("Imp",   "imp_calc",   "imp_meas"),
        ]:
            df_c[f"diff_{param}"] = df_c.apply(
                lambda r, c=c_col, m=m_col: pct_diff(r[c], r[m]) if r[m] and r[m] > 0 else np.nan,
                axis=1)

        st.markdown('<p class="sec">📋 FULL COMPARISON TABLE</p>', unsafe_allow_html=True)
        show = df_c[["calc_at","site_name",
                     "p_max_calc","p_max_meas","diff_Power",
                     "voc_calc","voc_meas","diff_Voc",
                     "isc_calc","isc_meas","diff_Isc"]].copy()
        show.columns = ["Timestamp","Site",
                        "P_calc (W)","P_meas (W)","ΔP (%)",
                        "Voc_calc (V)","Voc_meas (V)","ΔVoc (%)",
                        "Isc_calc (A)","Isc_meas (A)","ΔIsc (%)"]
        st.dataframe(show.style.format({
            "P_calc (W)":"{:.2f}","P_meas (W)":"{:.2f}","ΔP (%)":"{:.2f}",
            "Voc_calc (V)":"{:.3f}","Voc_meas (V)":"{:.3f}","ΔVoc (%)":"{:.2f}",
            "Isc_calc (A)":"{:.3f}","Isc_meas (A)":"{:.3f}","ΔIsc (%)":"{:.2f}",
        }), use_container_width=True, height=350)

        st.markdown('<p class="sec">📊 DEVIATION BAR CHARTS (%)</p>', unsafe_allow_html=True)
        diff_params = ["diff_Power","diff_Voc","diff_Vmp","diff_Isc","diff_Imp"]
        diff_labels = ["Power","Voc","Vmp","Isc","Imp"]
        fig = make_subplots(rows=2, cols=3,
            subplot_titles=diff_labels + [""],
            vertical_spacing=0.18)
        positions = [(1,1),(1,2),(1,3),(2,1),(2,2)]
        for dp, dl, pos in zip(diff_params, diff_labels, positions):
            valid = df_c[dp].dropna()
            if not valid.empty:
                colors = ["#00E676" if abs(v) <= 5 else "#FF4444" for v in valid]
                fig.add_trace(go.Bar(
                    x=list(range(len(valid))), y=valid.values, name=dl,
                    marker_color=colors,
                    text=[f"{v:.1f}%" for v in valid.values],
                    textposition="outside"),
                    row=pos[0], col=pos[1])
                fig.add_hline(y=5,  line=dict(color="#FFB800",dash="dash",width=1), row=pos[0], col=pos[1])
                fig.add_hline(y=-5, line=dict(color="#FFB800",dash="dash",width=1), row=pos[0], col=pos[1])
        fig.update_layout(height=500, showlegend=False,
            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,0.8)",
            font=dict(color="#8892A4",family="Exo 2"))
        for ann in fig.layout.annotations: ann.font.color = "#FFB800"
        fig.update_xaxes(gridcolor="rgba(255,184,0,.06)")
        fig.update_yaxes(gridcolor="rgba(255,184,0,.06)", title_text="Deviation (%)")
        st.plotly_chart(fig, use_container_width=True)

        st.markdown('<p class="sec">📈 PARITY PLOT — CALCULATED vs MEASURED</p>', unsafe_allow_html=True)
        valid_cmp = df_c[df_c.p_max_meas > 0].copy()
        if not valid_cmp.empty:
            mn = min(valid_cmp["p_max_calc"].min(), valid_cmp["p_max_meas"].min()) * 0.95
            mx = max(valid_cmp["p_max_calc"].max(), valid_cmp["p_max_meas"].max()) * 1.05
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(x=[mn,mx], y=[mn,mx],
                name="Ideal line (Calc = Meas)", line=dict(color="#FFB800",dash="dash",width=2)))
            fig2.add_trace(go.Scatter(
                x=valid_cmp["p_max_meas"], y=valid_cmp["p_max_calc"],
                mode="markers", name="Data points",
                marker=dict(color="#00D4FF", size=10, opacity=0.8)))
            fig2.update_layout(
                xaxis_title="Measured Power (W)", yaxis_title="Calculated Power / SEDA (W)",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,0.8)",
                font=dict(color="#8892A4",family="Exo 2"), height=420,
                title="Parity Plot — Points on the gold line indicate perfect agreement",
                title_font=dict(color="#FFB800"),
                legend=dict(bgcolor="rgba(0,0,0,.5)"))
            fig2.update_xaxes(gridcolor="rgba(255,184,0,.08)")
            fig2.update_yaxes(gridcolor="rgba(255,184,0,.08)")
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.warning("Enter measured values in the SEDA Calculator to populate the parity plot.")

        csv = io.StringIO(); df_c.to_csv(csv, index=False)
        st.download_button("⬇ Export Comparison CSV", csv.getvalue(),
                           "comparison_analysis.csv", "text/csv")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: IMPORT / EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📤 Import / Export":
    st.markdown('<p class="sec">📤 IMPORT / EXPORT DATA</p>', unsafe_allow_html=True)

    tab_imp, tab_exp = st.tabs(["📥 Import CSV / Excel", "📤 Export Data"])

    with tab_imp:
        template = pd.DataFrame({
            "recorded_at":   ["2024-01-01 08:00:00"],
            "site_name":     ["Default Site"],
            "irradiance":    [800.0],
            "pv_temp":       [45.0],
            "ambient_temp":  [30.0],
            "voc_measured":  [40.5],
            "vmp_measured":  [33.2],
            "isc_measured":  [10.05],
            "imp_measured":  [9.49],
            "power_measured":[315.0],
            "notes":         [""],
        })
        buf = io.BytesIO(); template.to_csv(buf, index=False)
        st.download_button("⬇ Download CSV Template", buf.getvalue(),
                           "template_field_measurement.csv", "text/csv")

        uploaded = st.file_uploader("Upload CSV or Excel file", type=["csv","xlsx","xls"])
        if uploaded:
            try:
                udf = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
                st.markdown(f"**Preview — {len(udf)} rows:**")
                st.dataframe(udf.head(10), use_container_width=True)
                required = ["recorded_at","irradiance","pv_temp","ambient_temp"]
                missing  = [c for c in required if c not in udf.columns]
                if missing:
                    st.error(f"Missing required columns: {missing}")
                else:
                    if "site_name" not in udf.columns:
                        udf["site_name"] = "Default Site"
                    if st.button("📥 Import Data", use_container_width=True):
                        with get_conn() as conn:
                            keep = [c for c in template.columns if c in udf.columns]
                            udf[keep].to_sql("measured_data", conn, if_exists="append", index=False)
                            conn.commit()
                        st.success(f"✅ {len(udf)} records imported successfully!")
            except Exception as e:
                st.error(f"Error reading file: {e}")

    with tab_exp:
        st.markdown('<p class="sec">📥 EXPORT TO EXCEL</p>', unsafe_allow_html=True)
        st.caption("Download a fully formatted Excel report with separate sheets for Field Measurements, Computation Results, and a Combined Comparison.")

        # ── Filter options ──────────────────────────────────────────────────────
        ex1, ex2 = st.columns(2)
        with ex1:
            export_site = st.selectbox("Filter by Site", ["All Sites"] + load_sites(), key="ex_site")
        with ex2:
            export_type = st.selectbox("Report Type", [
                "Full Report (All Sheets)",
                "Field Measurements Only",
                "Computation Results Only",
                "Combined Comparison Only",
            ], key="ex_type")

        def pct_diff_val(calc, meas):
            try:
                if meas and float(meas) != 0:
                    return round((float(calc) - float(meas)) / abs(float(meas)) * 100, 2)
            except: pass
            return None

        def build_excel(df_m, df_c, export_type):
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                wb = writer.book

                # ── Helper: style a worksheet ───────────────────────────────
                from openpyxl.styles import (PatternFill, Font, Alignment,
                                             Border, Side, GradientFill)
                from openpyxl.utils import get_column_letter

                GOLD    = "FFB800"
                NAVY    = "0A0E1A"
                PANEL   = "0D1525"
                CYAN    = "00D4FF"
                GREEN   = "00E676"
                RED     = "FF4444"
                WHITE   = "E8EAF0"
                MUTED   = "8892A4"
                ORANGE  = "FF6B00"

                def style_header_row(ws, row_num, num_cols, bg=NAVY, fg=GOLD, bold=True, size=11):
                    for col in range(1, num_cols + 1):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill      = PatternFill("solid", fgColor=bg)
                        cell.font      = Font(color=fg, bold=bold, size=size, name="Calibri")
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                def style_data_row(ws, row_num, num_cols, bg="111827", fg=WHITE):
                    for col in range(1, num_cols + 1):
                        cell = ws.cell(row=row_num, column=col)
                        cell.fill      = PatternFill("solid", fgColor=bg)
                        cell.font      = Font(color=fg, size=10, name="Calibri")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                def auto_col_width(ws):
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col:
                            try:
                                max_len = max(max_len, len(str(cell.value or "")))
                            except: pass
                        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 35)

                def add_title_block(ws, title, subtitle, num_cols):
                    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
                    t = ws.cell(row=1, column=1, value=f"☀ {title}")
                    t.fill      = PatternFill("solid", fgColor=NAVY)
                    t.font      = Font(color=GOLD, bold=True, size=16, name="Calibri")
                    t.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[1].height = 35

                    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
                    s = ws.cell(row=2, column=1, value=subtitle)
                    s.fill      = PatternFill("solid", fgColor=PANEL)
                    s.font      = Font(color=MUTED, size=10, name="Calibri")
                    s.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[2].height = 20

                    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
                    d = ws.cell(row=3, column=1,
                                value=f"Generated: {datetime.now().strftime('%d %B %Y  %H:%M:%S')}")
                    d.fill      = PatternFill("solid", fgColor=PANEL)
                    d.font      = Font(color=MUTED, size=9, italic=True, name="Calibri")
                    d.alignment = Alignment(horizontal="center", vertical="center")
                    ws.row_dimensions[3].height = 18

                # ── SHEET 1: Field Measurements ─────────────────────────────
                if export_type in ["Full Report (All Sheets)", "Field Measurements Only"]:
                    if not df_m.empty:
                        cols_meas = {
                            "id":             "ID",
                            "recorded_at":    "Timestamp",
                            "site_name":      "Site",
                            "irradiance":     "Irradiance\n(W/m²)",
                            "pv_temp":        "PV Temp\n(°C)",
                            "ambient_temp":   "Ambient Temp\n(°C)",
                            "voc_measured":   "Voc Measured\n(V)",
                            "vmp_measured":   "Vmp Measured\n(V)",
                            "isc_measured":   "Isc Measured\n(A)",
                            "imp_measured":   "Imp Measured\n(A)",
                            "power_measured": "Power Measured\n(W)",
                            "notes":          "Notes",
                        }
                        available = [c for c in cols_meas if c in df_m.columns]
                        df_out    = df_m[available].rename(columns=cols_meas)
                        df_out.to_excel(writer, sheet_name="Field Measurements",
                                        index=False, startrow=4)
                        ws = writer.sheets["Field Measurements"]
                        nc = len(df_out.columns)
                        add_title_block(ws,
                            "SOLAR PV — FIELD MEASUREMENTS",
                            "Real Operating Conditions (ROC) — Values measured at site", nc)
                        style_header_row(ws, 5, nc, bg=NAVY, fg=GOLD)
                        ws.row_dimensions[5].height = 30
                        for r in range(6, 6 + len(df_out)):
                            bg = "111827" if r % 2 == 0 else "0D1525"
                            style_data_row(ws, r, nc, bg=bg)
                        # Colour irradiance column gold
                        irr_col = list(df_out.columns).index("Irradiance\n(W/m²)") + 1
                        for r in range(6, 6 + len(df_out)):
                            ws.cell(row=r, column=irr_col).font = Font(color=GOLD, bold=True, size=10)
                        auto_col_width(ws)
                        ws.freeze_panes = "A6"

                # ── SHEET 2: Computation Results ────────────────────────────
                if export_type in ["Full Report (All Sheets)", "Computation Results Only"]:
                    if not df_c.empty:
                        cols_calc = {
                            "id":           "ID",
                            "calc_at":      "Computed At",
                            "site_name":    "Site",
                            "irradiance":   "Irradiance\n(W/m²)",
                            "t_mod":        "T_mod\n(°C)",
                            "dirt_pct":     "Dirt\n(%)",
                            "f_g":          "f_g",
                            "f_degrad":     "f_degrad",
                            "f_temp_p":     "f_temp_p",
                            "f_clean":      "f_clean",
                            "f_unshade":    "f_unshade",
                            "f_mm":         "f_mm",
                            "p_max_calc":   "P_max Calculated\n(W)",
                            "voc_calc":     "Voc Calculated\n(V)",
                            "vmp_calc":     "Vmp Calculated\n(V)",
                            "isc_calc":     "Isc Calculated\n(A)",
                            "imp_calc":     "Imp Calculated\n(A)",
                        }
                        available = [c for c in cols_calc if c in df_c.columns]
                        df_out    = df_c[available].rename(columns=cols_calc)
                        df_out.to_excel(writer, sheet_name="Computation Results",
                                        index=False, startrow=4)
                        ws = writer.sheets["Computation Results"]
                        nc = len(df_out.columns)
                        add_title_block(ws,
                            "SOLAR PV — SEDA COMPUTATION RESULTS",
                            "Estimated outputs based on SEDA Equations 5.7 – 5.14", nc)
                        style_header_row(ws, 5, nc, bg=NAVY, fg=ORANGE)
                        ws.row_dimensions[5].height = 30
                        for r in range(6, 6 + len(df_out)):
                            bg = "111827" if r % 2 == 0 else "0D1525"
                            style_data_row(ws, r, nc, bg=bg)
                        # Highlight P_max column
                        p_col = list(df_out.columns).index("P_max Calculated\n(W)") + 1
                        for r in range(6, 6 + len(df_out)):
                            ws.cell(row=r, column=p_col).font = Font(color=ORANGE, bold=True, size=10)
                        auto_col_width(ws)
                        ws.freeze_panes = "A6"

                # ── SHEET 3: Combined Comparison ────────────────────────────
                if export_type in ["Full Report (All Sheets)", "Combined Comparison Only"]:
                    if not df_c.empty:
                        rows = []
                        for _, row in df_c.iterrows():
                            d_p   = pct_diff_val(row.get("p_max_calc"),  row.get("p_max_meas"))
                            d_voc = pct_diff_val(row.get("voc_calc"),    row.get("voc_meas"))
                            d_vmp = pct_diff_val(row.get("vmp_calc"),    row.get("vmp_meas"))
                            d_isc = pct_diff_val(row.get("isc_calc"),    row.get("isc_meas"))
                            d_imp = pct_diff_val(row.get("imp_calc"),    row.get("imp_meas"))
                            rows.append({
                                "Timestamp":          row.get("calc_at",""),
                                "Site":               row.get("site_name",""),
                                "Irradiance (W/m²)":  row.get("irradiance",""),
                                "T_mod (°C)":         row.get("t_mod",""),
                                # Power
                                "P Calculated (W)":   row.get("p_max_calc",""),
                                "P Measured (W)":     row.get("p_max_meas",""),
                                "ΔP (%)":             d_p,
                                # Voc
                                "Voc Calc (V)":       row.get("voc_calc",""),
                                "Voc Meas (V)":       row.get("voc_meas",""),
                                "ΔVoc (%)":           d_voc,
                                # Vmp
                                "Vmp Calc (V)":       row.get("vmp_calc",""),
                                "Vmp Meas (V)":       row.get("vmp_meas",""),
                                "ΔVmp (%)":           d_vmp,
                                # Isc
                                "Isc Calc (A)":       row.get("isc_calc",""),
                                "Isc Meas (A)":       row.get("isc_meas",""),
                                "ΔIsc (%)":           d_isc,
                                # Imp
                                "Imp Calc (A)":       row.get("imp_calc",""),
                                "Imp Meas (A)":       row.get("imp_meas",""),
                                "ΔImp (%)":           d_imp,
                            })
                        df_cmp = pd.DataFrame(rows)
                        df_cmp.to_excel(writer, sheet_name="Comparison",
                                        index=False, startrow=4)
                        ws = writer.sheets["Comparison"]
                        nc = len(df_cmp.columns)
                        add_title_block(ws,
                            "SOLAR PV — CALCULATED vs MEASURED COMPARISON",
                            "Deviation analysis — Green = within ±5%  |  Red = exceeds ±5%", nc)
                        style_header_row(ws, 5, nc, bg=NAVY, fg=CYAN)
                        ws.row_dimensions[5].height = 30
                        for r_idx, r in enumerate(range(6, 6 + len(df_cmp))):
                            bg = "111827" if r_idx % 2 == 0 else "0D1525"
                            style_data_row(ws, r, nc, bg=bg)
                            # Colour ΔP, ΔVoc, ΔVmp, ΔIsc, ΔImp columns
                            for col_name, col_idx in [
                                ("ΔP (%)", 7), ("ΔVoc (%)", 10),
                                ("ΔVmp (%)", 13), ("ΔIsc (%)", 16), ("ΔImp (%)", 19)
                            ]:
                                cell = ws.cell(row=r, column=col_idx)
                                try:
                                    val = float(cell.value) if cell.value is not None else None
                                    if val is not None:
                                        color = GREEN if abs(val) <= 5 else RED
                                        cell.font = Font(color=color, bold=True, size=10)
                                except: pass
                        auto_col_width(ws)
                        ws.freeze_panes = "A6"

                # ── SHEET 4: Summary Stats ───────────────────────────────────
                if export_type == "Full Report (All Sheets)" and not df_m.empty:
                    num_cols_list = ["irradiance","pv_temp","ambient_temp",
                                     "voc_measured","vmp_measured",
                                     "isc_measured","imp_measured","power_measured"]
                    available = [c for c in num_cols_list if c in df_m.columns]
                    summary = df_m[available].describe().T.round(3)
                    summary.index.name = "Parameter"
                    summary = summary.reset_index()
                    summary["Parameter"] = summary["Parameter"].str.replace("_measured","").str.replace("_"," ").str.title()
                    summary.to_excel(writer, sheet_name="Summary Statistics",
                                     index=False, startrow=4)
                    ws = writer.sheets["Summary Statistics"]
                    nc = len(summary.columns)
                    add_title_block(ws,
                        "SOLAR PV — STATISTICAL SUMMARY",
                        "Descriptive statistics for all field measurement parameters", nc)
                    style_header_row(ws, 5, nc, bg=NAVY, fg=GREEN)
                    ws.row_dimensions[5].height = 25
                    for r in range(6, 6 + len(summary)):
                        bg = "111827" if r % 2 == 0 else "0D1525"
                        style_data_row(ws, r, nc, bg=bg)
                    auto_col_width(ws)

            buf.seek(0)
            return buf

        # ── Generate & Download button ──────────────────────────────────────
        st.markdown("---")
        if st.button("📊 GENERATE EXCEL REPORT", use_container_width=True):
            df_m_export = load_measured(
                None if export_site == "All Sites" else export_site)
            df_c_export = load_calcs(
                None if export_site == "All Sites" else export_site)

            if df_m_export.empty and df_c_export.empty:
                st.warning("No data found to export.")
            else:
                with st.spinner("Generating Excel report..."):
                    excel_buf = build_excel(df_m_export, df_c_export, export_type)
                fname = f"Solar_PV_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    label="⬇ DOWNLOAD EXCEL REPORT",
                    data=excel_buf,
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success(f"✅ Report ready — {fname}")


# ═══════════════════════════════════════════════════════════════════════════════
#  PAGE: ANALYTICS
# ═══════════════════════════════════════════════════════════════════════════════
elif page == "📈 Analytics":
    st.markdown('<p class="sec">📈 ADVANCED ANALYTICS</p>', unsafe_allow_html=True)

    tab1, tab2, tab3 = st.tabs(["🌞 Irradiance & Power", "🌡️ Temperature Effects", "📉 I-V Curve"])

    with tab1:
        if df_meas.empty:
            st.info("No field measurement data available.")
        else:
            fig = px.scatter(df_meas, x="irradiance", y="power_measured",
                color="pv_temp",
                color_continuous_scale=[[0,"#003B7A"],[.5,"#FF6B00"],[1,"#FF0000"]],
                title="Measured Power vs Solar Irradiance",
                labels={"irradiance":"Irradiance (W/m²)","power_measured":"Measured Power (W)","pv_temp":"PV Temp (°C)"},
                trendline="lowess")
            fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,0.8)",
                font=dict(color="#8892A4",family="Exo 2"), title_font=dict(color="#FFB800"))
            fig.update_xaxes(gridcolor="rgba(255,184,0,.08)")
            fig.update_yaxes(gridcolor="rgba(255,184,0,.08)")
            st.plotly_chart(fig, use_container_width=True)

            # Daily power statistics
            df_meas["date"] = df_meas["recorded_at"].dt.date
            daily = df_meas.groupby("date")["power_measured"].agg(["mean","max","min"]).reset_index()
            fig2 = go.Figure()
            fig2.add_trace(go.Bar(x=daily["date"], y=daily["mean"], name="Avg Power",
                marker_color="#FFB800", opacity=0.8))
            fig2.add_trace(go.Scatter(x=daily["date"], y=daily["max"], name="Max Power",
                line=dict(color="#FF6B00", width=2)))
            fig2.update_layout(height=320, paper_bgcolor="rgba(0,0,0,0)",
                plot_bgcolor="rgba(13,21,37,0.8)", font=dict(color="#8892A4",family="Exo 2"),
                title="Daily Power Statistics", title_font=dict(color="#FFB800"),
                legend=dict(bgcolor="rgba(0,0,0,.5)"))
            fig2.update_xaxes(gridcolor="rgba(255,184,0,.08)")
            fig2.update_yaxes(gridcolor="rgba(255,184,0,.08)", title="Power (W)")
            st.plotly_chart(fig2, use_container_width=True)

    with tab2:
        if df_meas.empty:
            st.info("No field measurement data available.")
        else:
            ca, cb = st.columns(2)
            with ca:
                fig = px.scatter(df_meas, x="pv_temp", y="voc_measured",
                    color="irradiance", color_continuous_scale="Plasma",
                    title="Voc vs PV Temperature", trendline="ols",
                    labels={"pv_temp":"PV Temp (°C)","voc_measured":"Voc (V)"})
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,.8)",
                    font=dict(color="#8892A4",family="Exo 2"), title_font=dict(color="#FFB800"))
                fig.update_xaxes(gridcolor="rgba(255,184,0,.08)")
                fig.update_yaxes(gridcolor="rgba(255,184,0,.08)")
                st.plotly_chart(fig, use_container_width=True)
            with cb:
                fig = px.scatter(df_meas, x="pv_temp", y="power_measured",
                    color="irradiance", color_continuous_scale="Hot",
                    title="Power vs PV Temperature", trendline="ols",
                    labels={"pv_temp":"PV Temp (°C)","power_measured":"Power (W)"})
                fig.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,.8)",
                    font=dict(color="#8892A4",family="Exo 2"), title_font=dict(color="#FFB800"))
                fig.update_xaxes(gridcolor="rgba(255,184,0,.08)")
                fig.update_yaxes(gridcolor="rgba(255,184,0,.08)")
                st.plotly_chart(fig, use_container_width=True)

            df_meas["temp_diff"] = df_meas["pv_temp"] - df_meas["ambient_temp"]
            fig3 = px.histogram(df_meas, x="temp_diff", nbins=30,
                title="Distribution: PV Temp − Ambient Temp (°C)",
                color_discrete_sequence=["#00D4FF"])
            fig3.update_layout(paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,.8)",
                font=dict(color="#8892A4",family="Exo 2"), title_font=dict(color="#FFB800"))
            fig3.update_xaxes(gridcolor="rgba(255,184,0,.08)", title="ΔT (°C)")
            fig3.update_yaxes(gridcolor="rgba(255,184,0,.08)")
            st.plotly_chart(fig3, use_container_width=True)

    with tab3:
        st.markdown("**Select a field record to view its approximate I-V & P-V curve:**")
        if df_meas.empty:
            st.info("No field measurement data available.")
        else:
            sel_id = st.selectbox("Record ID", df_meas["id"].tolist())
            rec    = df_meas[df_meas["id"] == sel_id].iloc[0]
            voc_r  = float(rec["voc_measured"] or 40)
            isc_r  = float(rec["isc_measured"] or 10)

            v_arr  = np.linspace(0, voc_r, 300)
            Vt     = 0.026 * 25
            i_arr  = isc_r * (1 - np.exp((v_arr - voc_r) / (Vt * 12)))
            i_arr  = np.clip(i_arr, 0, isc_r)
            p_arr  = v_arr * i_arr

            fig = make_subplots(specs=[[{"secondary_y": True}]])
            fig.add_trace(go.Scatter(x=v_arr, y=i_arr, name="Current (A)",
                line=dict(color="#00D4FF", width=2.5)), secondary_y=False)
            fig.add_trace(go.Scatter(x=v_arr, y=p_arr, name="Power (W)",
                line=dict(color="#FFB800", width=2, dash="dash")), secondary_y=True)
            mpp = np.argmax(p_arr)
            fig.add_trace(go.Scatter(x=[v_arr[mpp]], y=[i_arr[mpp]], name="MPP",
                mode="markers", marker=dict(color="#FF6B00", size=14, symbol="star")),
                secondary_y=False)
            fig.update_layout(
                title=f"I-V & P-V Curve — Record #{int(sel_id)} | Irr={rec['irradiance']:.0f} W/m² | T={rec['pv_temp']:.1f}°C",
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(13,21,37,.8)",
                font=dict(color="#8892A4",family="Exo 2"), title_font=dict(color="#FFB800"),
                legend=dict(bgcolor="rgba(0,0,0,.5)"))
            fig.update_xaxes(title_text="Voltage (V)", gridcolor="rgba(255,184,0,.08)")
            fig.update_yaxes(title_text="Current (A)", secondary_y=False, gridcolor="rgba(255,184,0,.08)")
            fig.update_yaxes(title_text="Power (W)",   secondary_y=True)
            st.plotly_chart(fig, use_container_width=True)

            c1,c2,c3,c4 = st.columns(4)
            c1.metric("Voc (measured)",  f"{rec['voc_measured']:.2f} V")
            c2.metric("Vmp (measured)",  f"{rec['vmp_measured']:.2f} V")
            c3.metric("Isc (measured)",  f"{rec['isc_measured']:.3f} A")
            c4.metric("Imp (measured)",  f"{rec['imp_measured']:.3f} A")

            # Statistical summary
            st.markdown("---")
            st.markdown('<p class="sec">📋 STATISTICAL SUMMARY</p>', unsafe_allow_html=True)
            num_cols = ["irradiance","pv_temp","ambient_temp",
                        "voc_measured","vmp_measured","isc_measured","imp_measured","power_measured"]
            summary = df_meas[num_cols].describe().T
            summary.index.name = "Parameter"
            st.dataframe(summary.style.format("{:.3f}"), use_container_width=True)
